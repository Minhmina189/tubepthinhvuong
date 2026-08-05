# -*- coding: utf-8 -*-
"""
Tạo file báo giá Excel theo định dạng Nam An.
"""
import math
import re
import unicodedata
from datetime import date
from pathlib import Path
from typing import List, Optional

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from product_matcher import MatchResult


# ─────────────────────────────────────────────
# STYLE CONSTANTS
# ─────────────────────────────────────────────
COLOR_HEADER_BG  = "1F4E79"   # Dark blue header
COLOR_COL_TITLE  = "2E75B6"   # Column title blue
COLOR_ALT_ROW    = "EBF3FB"   # Light blue alt row
COLOR_YELLOW     = "FFFF00"
COLOR_WHITE      = "FFFFFF"

FONT_TITLE  = Font(name="Times New Roman", bold=True, size=14, color=COLOR_WHITE)
FONT_HEADER = Font(name="Times New Roman", bold=True, size=10, color=COLOR_WHITE)
FONT_BODY   = Font(name="Times New Roman", size=10)
FONT_BODY_B = Font(name="Times New Roman", bold=True, size=10)
FONT_SMALL  = Font(name="Times New Roman", size=9)

THIN = Side(style="thin", color="000000")
MED  = Side(style="medium", color="000000")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_MED = Border(left=MED,  right=MED,  top=MED,  bottom=MED)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=True)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _row_height_for(*text_col_pairs, min_h: float = 20) -> float:
    """Ước tính chiều cao hàng (pts) dựa trên nội dung dài nhất."""
    max_lines = 1
    for text, col_w in text_col_pairs:
        if text:
            chars_per_line = max(1, int(col_w * 0.95))
            max_lines = max(max_lines, math.ceil(len(str(text)) / chars_per_line))
    return max(min_h, max_lines * 14.4)


def _cell(ws, row: int, col: int, value=None, font=None, fill=None,
          alignment=None, border=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:        c.font        = font
    if fill:        c.fill        = fill
    if alignment:   c.alignment   = alignment
    if border:      c.border      = border
    if number_format: c.number_format = number_format
    return c


def _merge(ws, row1: int, col1: int, row2: int, col2: int,
           value=None, font=None, fill=None, alignment=None,
           border=None, number_format=None):
    """Merge cells, set value/style, and apply borders to ALL cells in the range."""
    ws.merge_cells(
        f"{get_column_letter(col1)}{row1}:{get_column_letter(col2)}{row2}"
    )
    c = ws.cell(row=row1, column=col1, value=value)
    if font:         c.font        = font
    if alignment:    c.alignment   = alignment
    if number_format: c.number_format = number_format

    # Fill every cell in the range (not just top-left)
    if fill:
        for r in range(row1, row2 + 1):
            for col in range(col1, col2 + 1):
                ws.cell(row=r, column=col).fill = fill

    # Apply border to every cell with correct sides (outer edges only)
    if border:
        for r in range(row1, row2 + 1):
            for col in range(col1, col2 + 1):
                ws.cell(row=r, column=col).border = Border(
                    left   = border.left   if col == col1 else Side(style=None),
                    right  = border.right  if col == col2 else Side(style=None),
                    top    = border.top    if r   == row1 else Side(style=None),
                    bottom = border.bottom if r   == row2 else Side(style=None),
                )
    return c


# ─────────────────────────────────────────────
# INPUT PARSING
# ─────────────────────────────────────────────

def parse_customer_excel(filepath: str) -> List[dict]:
    """
    Đọc file Excel / CSV đơn hàng của khách.
    Quét toàn bộ cột để nhận diện thông tin — hỗ trợ các định dạng:
      • Danh mục hàng hóa | Ký mã hiệu | Tiêu chuẩn | Cấu hình kỹ thuật | ĐVT | Khối lượng
      • STT | Tên hàng / Sản phẩm | Chất liệu | Số lượng | Ghi chú
    Trả về list[dict] với keys: name, material, qty, note_in, raw_desc
    """
    ext = Path(filepath).suffix.lower()
    if ext == ".csv":
        import csv
        rows = []
        with open(filepath, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(_parse_row(row))
        return [r for r in rows if r]

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # ── Phát hiện hàng tiêu đề ─────────────────────────────────────────
    # Tìm hàng có nhiều cột nhất + có từ khóa sản phẩm/số lượng
    header_row = 1
    headers: dict = {}
    best_score = 0

    for r in range(1, min(8, ws.max_row + 1)):
        cells = [str(ws.cell(r, c).value or "").strip()
                 for c in range(1, min(ws.max_column + 1, 25))]
        row_text = " ".join(c.lower() for c in cells if c)
        non_empty = sum(1 for c in cells if c)
        score = non_empty
        if any(k in row_text for k in ["tên","hàng","sản phẩm","danh mục","product","item"]):
            score += 5
        if any(k in row_text for k in ["số lượng","khối lượng","quantity","qty","sl","kl"]):
            score += 3
        if score > best_score and non_empty >= 2:
            best_score = score
            header_row = r
            headers = {val.lower(): i + 1 for i, val in enumerate(cells) if val}

    # ── Nhận diện cột ──────────────────────────────────────────────────
    col_name = _find_col(headers, [
        "tên hàng","sản phẩm","san pham","ten hang","product","item",
        "hàng hóa","danh mục hàng hóa","danh mục","danh muc",
        "mô tả","mo ta","tên sp","ten sp","tên sản phẩm","hàng","hang",
    ])
    # Cột ký mã hiệu / model code (chứa size + cấp độ như "M14x110 (10.9)")
    col_spec = _find_col(headers, [
        "ký mã hiệu","ky ma hieu","mã hiệu","ma hieu","model","spec",
        "ký hiệu","ky hieu","kỹ mã hiệu","ma sp","ký mã",
    ])
    # Cột mô tả kỹ thuật (chứa vật liệu, thành phần bộ)
    col_desc = _find_col(headers, [
        "cấu hình, tính năng kỹ thuật cơ bản","cấu hình","cau hinh",
        "tính năng","tinh nang","mô tả kỹ thuật","mo ta ky thuat",
        "description","chi tiết","chi tiet","đặc tính","thông số",
    ])
    col_mat  = _find_col(headers, [
        "chất liệu","chat lieu","vật liệu","material","mat","loại",
    ])
    col_qty  = _find_col(headers, [
        "số lượng","so luong","quantity","qty","sl","số l",
        "khối lượng","khoi luong","kl","số lượng yêu cầu","slg",
    ])
    col_unit = _find_col(headers, [
        "đơn vị","don vi","dvt","đvt","unit","đơn vị tính","don vi tinh",
    ])
    col_note = _find_col(headers, [
        "ghi chú","ghi chu","note","notes","chú thích","chu thich",
    ])

    results = []
    for r in range(header_row + 1, ws.max_row + 1):
        def v(col, _r=r):
            if col is None: return ""
            return str(ws.cell(_r, col).value or "").strip()

        name = v(col_name)
        spec = v(col_spec)   # "M14x110 (10.9)"
        desc = v(col_desc)   # "M14x110, loại 10.9, 2 đệm phẳng, ê cu, bulong đen"
        unit = v(col_unit).lower() if col_unit else ""

        # Nếu không có cột tên, thử dùng spec
        if not name and spec:
            name = spec
        if not name:
            continue

        # Làm giàu tên bằng thông tin từ ký mã hiệu (spec)
        enriched_name = name
        if spec:
            name_has_size = bool(re.search(r'M\d+', name, re.IGNORECASE))
            spec_has_size = bool(re.search(r'M\d+', spec, re.IGNORECASE))
            if not name_has_size and spec_has_size:
                # Tên thiếu kích thước → dùng spec để bổ sung (vd "Bulong van chặn" + "M18x100 (8.8)")
                enriched_name = f"{name} {spec}"
            else:
                # Tên đã có kích thước → chỉ thêm cấp độ từ spec nếu chưa có
                grade_m = re.search(r'\((\d+\.?\d*)\)', spec)
                if grade_m and grade_m.group(1) in ("8.8","10.9","12.9"):
                    if grade_m.group(1) not in enriched_name:
                        enriched_name = f"{name} {grade_m.group(1)}"

        # Chất liệu: cột riêng → mô tả kỹ thuật → ký mã hiệu
        mat_str = v(col_mat)
        if not mat_str:
            mat_str = _extract_material_hint(f"{desc} {spec}")

        # Số lượng
        qty_str = v(col_qty)
        try:
            qty = int(float(qty_str)) if qty_str else 1
        except ValueError:
            qty = 1

        # raw_desc gửi sang matcher để phát hiện bộ (set):
        # thêm "bộ" prefix nếu đơn vị là Bộ — đây là tín hiệu mạnh nhất
        raw_desc_full = (f"[bộ] {desc}" if unit in ("bộ","bo","set") else desc).strip()

        # Ghi chú (kết hợp note + desc để hiển thị trong form)
        note_parts = [p for p in [v(col_note), desc] if p]
        note_in = " | ".join(note_parts)

        results.append({
            "name":     enriched_name,
            "material": mat_str,
            "qty":      max(1, qty),
            "note_in":  note_in,
            "raw_desc": raw_desc_full,
        })

    return results


def _find_col(headers: dict, candidates: List[str]) -> Optional[int]:
    # Normalize header keys to NFC to handle Excel Unicode encoding differences
    norm_h = {unicodedata.normalize('NFC', k): v for k, v in headers.items()}
    # Exact match first
    for c in candidates:
        nc = unicodedata.normalize('NFC', c)
        if nc in norm_h:
            return norm_h[nc]
    # Substring: candidate is contained in a header key (handles long compound headers)
    for c in candidates:
        nc = unicodedata.normalize('NFC', c)
        for h_key, col in norm_h.items():
            if nc in h_key or h_key.startswith(nc):
                return col
    return None


def _extract_material_hint(text: str) -> str:
    """Trích xuất gợi ý chất liệu từ cột mô tả/ký mã hiệu.
    Ưu tiên: mạ kẽm > inox > grade đen."""
    if not text:
        return ""
    # Normalize: NFC + chuyển dấu phẩy thập phân → chấm
    t = unicodedata.normalize('NFC', text.lower())
    t = re.sub(r'(\d),(\d)', r'\1.\2', t)

    # Mạ kẽm (zinc plating) — "mã kẽm" là typo phổ biến (ã thay ạ)
    if any(k in t for k in ["mạ kẽm", "mã kẽm", "ma kẽm", "ma kem", "galvanized", "hdg"]):
        if "10.9" in t: return "Thép 10.9 mạ kẽm"
        if "12.9" in t: return "Thép 12.9 mạ kẽm"
        return "Thép mạ kẽm 8.8"
    if any(k in t for k in ["inox", "sus", "stainless", "không gỉ"]):
        if "316" in t: return "Inox 316"
        if "201" in t: return "Inox 201"
        return "Inox 304"
    if "12.9" in t: return "Thép đen 12.9"
    if "10.9" in t: return "Thép đen 10.9"
    if "8.8"  in t: return "Thép đen 8.8"
    if re.search(r'\bmạ\b', t): return "Thép mạ kẽm 8.8"
    return ""


def _parse_row(row: dict) -> Optional[dict]:
    name = ""
    for k in row:
        if any(x in k.lower() for x in ["tên","san pham","product","item","hàng"]):
            name = str(row[k] or "").strip()
    if not name:
        return None
    mat = qty_s = note = ""
    for k, v in row.items():
        kl = k.lower()
        if any(x in kl for x in ["chất","material","vật liệu"]):
            mat = str(v or "")
        elif any(x in kl for x in ["số lượng","qty","quantity"]):
            qty_s = str(v or "")
        elif any(x in kl for x in ["ghi","note"]):
            note = str(v or "")
    try:
        qty = int(float(qty_s)) if qty_s else 1
    except ValueError:
        qty = 1
    return {"name": name, "material": mat, "qty": max(1, qty), "note_in": note}


# ─────────────────────────────────────────────
# QUOTE DATA
# ─────────────────────────────────────────────

class QuoteItem:
    def __init__(self, stt: int, raw_name: str, raw_mat: str, qty: int,
                 result: MatchResult, note_in: str = ""):
        self.stt       = stt
        self.raw_name  = raw_name
        self.raw_mat   = raw_mat
        self.qty       = qty
        self.result    = result
        self.note_in   = note_in

    @property
    def product_name(self) -> str:
        return self.result.product_name or self.raw_name

    @property
    def material(self) -> str:
        return self.result.material_display or self.raw_mat

    @property
    def unit_price(self) -> Optional[int]:
        return self.result.unit_price

    @property
    def total(self) -> Optional[int]:
        if self.unit_price is not None:
            return self.unit_price * self.qty
        return None

    @property
    def note(self) -> str:
        parts = []
        if self.note_in:
            parts.append(self.note_in)
        if self.result.note:
            parts.append(self.result.note)
        if self.unit_price is None:
            parts.append("⚠ Không tìm được giá")
        return " | ".join(parts)


# ─────────────────────────────────────────────
# EXCEL WRITER
# ─────────────────────────────────────────────

def generate_quote_excel(
    items: List[QuoteItem],
    output_path: str,
    quote_number: str = "",
    receiver_name: str = "",
    receiver_address: str = "",
    receiver_tel: str = "",
    quote_date: Optional[date] = None,
    vat_rate: int = 8,
) -> str:
    """
    Tạo file Excel báo giá theo định dạng Nam An.
    Trả về đường dẫn file đã tạo.
    """
    if quote_date is None:
        quote_date = date.today()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Báo Giá"

    # Column widths
    col_widths = [6, 42, 18, 8, 10, 13, 16, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── LOGO / COMPANY HEADER ─────────────────
    ws.merge_cells(f"A{row}:H{row}")
    _cell(ws, row, 1, "CÔNG TY TNHH VẬT TƯ CÔNG NGHIỆP NAM AN",
          font=Font(name="Times New Roman", bold=True, size=13, color=COLOR_HEADER_BG),
          fill=_fill(COLOR_WHITE), alignment=ALIGN_CENTER)
    row += 1

    ws.merge_cells(f"A{row}:H{row}")
    _cell(ws, row, 1,
          "Tổng kho/VPGD: Số 68 khu đất DV Vân Canh, KĐT mới Vân Canh, Hoài Đức, Hà Nội",
          font=FONT_SMALL, alignment=ALIGN_CENTER)
    row += 1

    ws.merge_cells(f"A{row}:H{row}")
    _cell(ws, row, 1,
          "Tel: 098 334 41 63 – 098 350 63 89 | Email: bulongocvitnaman@gmail.com | Web: bulongnaman.vn",
          font=FONT_SMALL, alignment=ALIGN_CENTER)
    row += 1
    row += 1  # spacer

    # ── TITLE ─────────────────────────────────
    ws.merge_cells(f"A{row}:H{row}")
    ws.row_dimensions[row].height = 28
    _cell(ws, row, 1, "THƯ BÁO GIÁ",
          font=Font(name="Times New Roman", bold=True, size=16),
          alignment=ALIGN_CENTER)
    row += 1

    if quote_number:
        ws.merge_cells(f"A{row}:H{row}")
        _cell(ws, row, 1, quote_number,
              font=Font(name="Times New Roman", size=11),
              alignment=ALIGN_CENTER)
        row += 1

    # Date
    ws.merge_cells(f"A{row}:H{row}")
    _cell(ws, row, 1,
          f"Ngày {quote_date.day:02d} tháng {quote_date.month:02d} năm {quote_date.year}",
          font=FONT_SMALL, alignment=Alignment(horizontal="right", vertical="center"))
    row += 1
    row += 1

    # ── PARTIES ───────────────────────────────
    party_fill = _fill("D9E1F2")

    # Header row: "Bên giao" | "Bên nhận"
    ws.row_dimensions[row].height = 18
    _merge(ws, row, 1, row, 4, "Bên giao",
           font=FONT_BODY_B, fill=party_fill, border=BORDER_ALL, alignment=ALIGN_CENTER)
    _merge(ws, row, 5, row, 8, "Bên nhận",
           font=FONT_BODY_B, fill=party_fill, border=BORDER_ALL, alignment=ALIGN_CENTER)
    row += 1

    # Data rows: label (A:B) | value (C:D) || label (E:F) | value (G:H)
    party_rows = [
        ("Đại diện", "Lê Ngọc Dũng",
         "Tên",     receiver_name     or ""),
        ("Địa chỉ",  "Số 68 khu đất DV Vân Canh, KĐT mới Vân Canh, Hoài Đức, Hà Nội",
         "Địa chỉ", receiver_address  or ""),
        ("Tel",      "",
         "Tel",     receiver_tel      or ""),
    ]
    for label_l, val_l, label_r, val_r in party_rows:
        ws.row_dimensions[row].height = 30
        _merge(ws, row, 1, row, 2, label_l, font=FONT_BODY_B, border=BORDER_ALL, alignment=ALIGN_LEFT)
        _merge(ws, row, 3, row, 4, val_l,   font=FONT_BODY,   border=BORDER_ALL,
               alignment=Alignment(horizontal="left", vertical="center", wrap_text=True))
        _merge(ws, row, 5, row, 6, label_r, font=FONT_BODY_B, border=BORDER_ALL, alignment=ALIGN_LEFT)
        _merge(ws, row, 7, row, 8, val_r,   font=FONT_BODY,   border=BORDER_ALL,
               alignment=Alignment(horizontal="left", vertical="center", wrap_text=True))
        row += 1

    row += 1

    # ── INTRO TEXT ────────────────────────────
    ws.merge_cells(f"A{row}:H{row}")
    _cell(ws, row, 1,
          "Công ty TNHH Vật Tư Công Nghiệp Nam An xin gửi đến quý công ty lời chào trân trọng nhất. "
          "Theo như đơn đặt hàng đã được phê duyệt, hai bên tiến hành giao nhận số lượng hàng hóa như sau:",
          font=FONT_BODY, alignment=Alignment(horizontal="left", vertical="center", wrap_text=True))
    ws.row_dimensions[row].height = 30
    row += 1
    row += 1

    # ── TABLE HEADER ──────────────────────────
    col_titles = ["STT", "Sản phẩm", "Chất liệu", "Đơn vị", "Số lượng", "Đơn giá", "Thành tiền", "Ghi chú"]
    header_fill = _fill(COLOR_COL_TITLE)
    ws.row_dimensions[row].height = 30
    for ci, title in enumerate(col_titles, 1):
        _cell(ws, row, ci, title,
              font=FONT_HEADER, fill=header_fill,
              border=BORDER_ALL, alignment=ALIGN_CENTER)
    table_header_row = row
    row += 1

    # ── TABLE ROWS ────────────────────────────
    _ = row  # first_data_row anchor (unused but marks table start)
    vat_base = 0
    for item in items:
        alt = (item.stt % 2 == 0)
        row_fill = _fill(COLOR_ALT_ROW) if alt else _fill(COLOR_WHITE)
        ws.row_dimensions[row].height = _row_height_for(
            (item.product_name, 42),
            (item.material, 18),
            (item.note or "", 22),
        )

        # Highlight rows with missing price
        if item.unit_price is None:
            row_fill = _fill("FFF2CC")  # yellow warning

        _cell(ws, row, 1, item.stt,      font=FONT_BODY, fill=row_fill, border=BORDER_ALL, alignment=ALIGN_CENTER)
        _cell(ws, row, 2, item.product_name, font=FONT_BODY, fill=row_fill, border=BORDER_ALL, alignment=ALIGN_LEFT)
        _cell(ws, row, 3, item.material, font=FONT_BODY, fill=row_fill, border=BORDER_ALL, alignment=ALIGN_LEFT)
        _cell(ws, row, 4, "Cái",         font=FONT_BODY, fill=row_fill, border=BORDER_ALL, alignment=ALIGN_CENTER)
        _cell(ws, row, 5, item.qty,      font=FONT_BODY, fill=row_fill, border=BORDER_ALL, alignment=ALIGN_CENTER)
        _cell(ws, row, 6, item.unit_price,
              font=FONT_BODY, fill=row_fill, border=BORDER_ALL, alignment=ALIGN_RIGHT,
              number_format="#,##0")
        _cell(ws, row, 7, item.total,
              font=FONT_BODY, fill=row_fill, border=BORDER_ALL, alignment=ALIGN_RIGHT,
              number_format="#,##0")
        _cell(ws, row, 8, item.note or "",
              font=FONT_BODY, fill=row_fill, border=BORDER_ALL, alignment=ALIGN_LEFT)

        if item.total:
            vat_base += item.total
        row += 1

    # ── TOTALS ────────────────────────────────
    row += 1
    vat_amount = round(vat_base * vat_rate / 100)
    total_with_vat = vat_base + vat_amount

    total_rows = [
        ("Tổng chưa VAT",           vat_base),
        (f"VAT ({vat_rate}%)",       vat_amount),
        ("Tổng gồm VAT",             total_with_vat),
    ]
    tot_fill = _fill("D9E1F2")
    for label, val in total_rows:
        _merge(ws, row, 1, row, 6, label,
               font=FONT_BODY_B, fill=tot_fill, border=BORDER_ALL, alignment=ALIGN_RIGHT)
        _cell(ws, row, 7, val,
              font=FONT_BODY_B, fill=tot_fill, border=BORDER_ALL, alignment=ALIGN_RIGHT,
              number_format="#,##0")
        _cell(ws, row, 8, "VND",
              font=FONT_BODY_B, fill=tot_fill, border=BORDER_ALL, alignment=ALIGN_CENTER)
        row += 1

    # ── CONDITIONS ────────────────────────────
    row += 1
    conditions = [
        "Tình trạng hàng hóa: Chưa qua sử dụng, đầy đủ phụ kiện",
        "Thanh toán: Thanh toán 100% trước khi nhận hàng.",
        "Thời gian giao hàng: 20-25 ngày kể từ khi có xác nhận đơn hàng.",
        "Thời gian bảo hành: 30 ngày khi có lỗi từ bên cung cấp.",
        "Giao hàng tại: Tại kho Nam An",
    ]
    for cond in conditions:
        ws.merge_cells(f"A{row}:H{row}")
        _cell(ws, row, 1, cond, font=FONT_BODY, alignment=ALIGN_LEFT)
        row += 1

    row += 1

    # ── SIGNATURES ────────────────────────────
    ws.merge_cells(f"A{row}:D{row}")
    _cell(ws, row, 1, "Đại diện kho bên nhận", font=FONT_BODY_B, alignment=ALIGN_CENTER)
    ws.merge_cells(f"E{row}:H{row}")
    _cell(ws, row, 5, "Đại diện bên tiếp nhận", font=FONT_BODY_B, alignment=ALIGN_CENTER)
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    _cell(ws, row, 1, "Công Ty TNHH Vật Tư Công Nghiệp Nam An", font=FONT_BODY_B, alignment=ALIGN_CENTER)
    ws.merge_cells(f"E{row}:H{row}")
    _cell(ws, row, 5, receiver_name or "", font=FONT_BODY_B, alignment=ALIGN_CENTER)
    row += 3  # space for signature

    ws.merge_cells(f"A{row}:D{row}")
    _cell(ws, row, 1, "Lê Ngọc Dũng", font=FONT_BODY_B, alignment=ALIGN_CENTER)

    # ── FREEZE PANES ──────────────────────────
    ws.freeze_panes = f"A{table_header_row + 1}"

    # ── PRINT SETTINGS ───────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = f"{table_header_row}:{table_header_row}"
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5

    wb.save(output_path)
    return output_path


# ─────────────────────────────────────────────
# PARSE INPUT EXCEL & MATCH PRICES
# ─────────────────────────────────────────────

def process_customer_file(
    input_path: str,
    output_path: str,
    quote_number: str = "",
    receiver_name: str = "",
    receiver_address: str = "",
    receiver_tel: str = "",
) -> dict:
    """
    Full pipeline: đọc file khách → match sản phẩm → tạo báo giá.
    Trả về dict với thống kê.
    """
    from product_matcher import match_product

    raw_items = parse_customer_excel(input_path)
    if not raw_items:
        return {"error": "Không đọc được dữ liệu từ file. Kiểm tra định dạng cột."}

    quote_items = []
    matched = 0
    missing_price = 0

    for i, raw in enumerate(raw_items, 1):
        result = match_product(raw["name"], raw["material"], raw.get("raw_desc", ""))
        item = QuoteItem(
            stt=i,
            raw_name=raw["name"],
            raw_mat=raw["material"],
            qty=raw["qty"],
            result=result,
            note_in=raw.get("note_in", ""),
        )
        quote_items.append(item)
        if result.matched:
            matched += 1
        if result.unit_price is None:
            missing_price += 1

    generate_quote_excel(
        items=quote_items,
        output_path=output_path,
        quote_number=quote_number,
        receiver_name=receiver_name,
        receiver_address=receiver_address,
        receiver_tel=receiver_tel,
    )

    total_val = sum(it.total or 0 for it in quote_items)
    return {
        "total_items":   len(quote_items),
        "matched":       matched,
        "missing_price": missing_price,
        "total_value":   total_val,
        "output_path":   output_path,
    }
